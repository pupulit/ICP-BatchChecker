# -*- coding: utf-8 -*-
# 二开：https://github.com/wongzeon/ICP-Checker
import re
import os
import cv2
import time
import base64
import hashlib
import time
import requests
import openpyxl as xl
from openpyxl.styles import Alignment
import argparse
import datetime

os.environ['no_proxy'] = '*'


def query_base(testsss):
    print("[*]查询数据："+testsss)
    try:
        info = testsss
        # 过滤空值和特殊字符，只允许 - . () 分别用于域名和公司名
        if info == "":
            raise ValueError("InputNone")
        info = re.sub("[^\\u4e00-\\u9fa5-A-Za-z0-9,-.()（）]", "", info)
        input_zh = re.compile(u'[\u4e00-\u9fa5]')
        zh_match = input_zh.search(info)
        if zh_match:
            info_result = info
        else:
            # 检测是否为可备案的域名类型（类型同步日期2022/01/06）
            input_url = re.compile(r'([^.]+)(?:\.(?:GOV\.cn|ORG\.cn|AC\.cn|MIL\.cn|NET\.cn|EDU\.cn|COM\.cn|BJ\.cn|TJ\.cn|SH\.cn|CQ\.cn|HE\.cn|SX\.cn|NM\.cn|LN\.cn|JL\.cn|HL\.cn|JS\.cn|ZJ\.cn|AH\.cn|FJ\.cn|JX\.cn|SD\.cn|HA\.cn|HB\.cn|HN\.cn|GD\.cn|GX\.cn|HI\.cn|SC\.cn|GZ\.cn|YN\.cn|XZ\.cn|SN\.cn|GS\.cn|QH\.cn|NX\.cn|XJ\.cn|TW\.cn|HK\.cn|MO\.cn|cn|REN|WANG|CITIC|TOP|SOHU|XIN|COM|NET|CLUB|XYZ|VIP|SITE|SHOP|INK|INFO|MOBI|RED|PRO|KIM|LTD|GROUP|BIZ|AUTO|LINK|WORK|LAW|BEER|STORE|TECH|FUN|ONLINE|ART|DESIGN|WIKI|LOVE|CENTER|VIDEO|SOCIAL|TEAM|SHOW|COOL|ZONE|WORLD|TODAY|CITY|CHAT|COMPANY|LIVE|FUND|GOLD|PLUS|GURU|RUN|PUB|EMAIL|LIFE|CO|FASHION|FIT|LUXE|YOGA|BAIDU|CLOUD|HOST|SPACE|PRESS|WEBSITE|ARCHI|ASIA|BIO|BLACK|BLUE|GREEN|LOTTO|ORGANIC|PET|PINK|POKER|PROMO|SKI|VOTE|VOTO|ICU))', flags=re.IGNORECASE)
            info_result = input_url.search(info)
            if info_result is None:
                if info.split(".")[0] == "":
                    raise ValueError("OnlyDomainInput")
                raise ValueError("ValidType")
            else:
                info_result = info_result.group()
        info_data = {'pageNum': '1', 'pageSize': '40', 'unitName': info_result, "serviceType":"1"}
        return info_data
    except ValueError as e:
        if str(e) == 'InputNone' or str(e) == 'OnlyDomainInput':
            print("\n ************** 输入内容有误 **************\n")
        else:
            print("\n*** 该域名不支持备案，请查阅：http://xn--fiq8ituh5mn9d1qbc28lu5dusc.xn--vuq861b/ ***\n")


def get_cookies():
    cookie_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32'}
    err_num = 0
    while err_num < 3:
        try:
            cookie = requests.utils.dict_from_cookiejar(requests.get('https://beian.miit.gov.cn/', headers=cookie_headers).cookies)['__jsluid_s']
            return cookie
        except:
            err_num += 1
            time.sleep(3)
    return -1


def get_token():
    timeStamp = round(time.time()*1000)
    authSecret = 'testtest' + str(timeStamp)
    authKey = hashlib.md5(authSecret.encode(encoding='UTF-8')).hexdigest()
    auth_data = {'authKey': authKey, 'timeStamp': timeStamp}
    url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/auth'
    try:
        t_response = requests.post(url=url, data=auth_data, headers=base_header).json()
        token = t_response['params']['bussiness']
    except:
        return -1
    return token


def get_check_pic(token):
    url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/image/getCheckImage'
    base_header['Accept'] = 'application/json, text/plain, */*'
    base_header.update({'Content-Length': '0', 'token': token})
    try:
        p_request = requests.post(url=url, data='', headers=base_header).json()
        p_uuid = p_request['params']['uuid']
        big_image = p_request['params']['bigImage']
        small_image = p_request['params']['smallImage']
    except:
        return -1
    # 解码图片，写入并计算图片缺口位置
    with open('bigImage.jpg', 'wb') as f:
        f.write(base64.b64decode(big_image))
    with open('smallImage.jpg', 'wb') as f:
        f.write(base64.b64decode(small_image))
    background_image = cv2.imread('bigImage.jpg', cv2.COLOR_GRAY2RGB)
    fill_image = cv2.imread('smallImage.jpg', cv2.COLOR_GRAY2RGB)
    position_match = cv2.matchTemplate(background_image, fill_image, cv2.TM_CCOEFF_NORMED)
    max_loc = cv2.minMaxLoc(position_match)[3][0]
    mouse_length = max_loc+1
    os.remove('bigImage.jpg')
    os.remove('smallImage.jpg')
    check_data = {'key': p_uuid, 'value': mouse_length}
    return check_data


def get_sign(check_data, token):
    check_url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/image/checkImage'
    base_header.update({'Content-Length': '60', 'token': token, 'Content-Type': 'application/json'})
    try:
        pic_sign = requests.post(check_url, json=check_data, headers=base_header).json()
        sign = pic_sign['params']
    except:
        return -1
    return sign


def get_beian_info(info_data, p_uuid, token, sign):
    domain_list = []
    info_url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/icpAbbreviateInfo/queryByCondition'
    base_header.update({'Content-Length': '78', 'uuid': p_uuid, 'token': token, 'sign': sign})
    try:
        beian_info = requests.post(url=info_url, json=info_data, headers=base_header).json()
        domain_total = beian_info['params']['total']
        page_total = beian_info['params']['lastPage']
        end_row = beian_info['params']['endRow']
        info = info_data['unitName']
        print(f"[*]查询对象：{info} > 共有 {domain_total} 个已备案域名")
        for i in range(0, page_total):
            # print(f"正在查询第{i+1}页......")
            for k in range(0, end_row+1):
                info_base = beian_info['params']['list'][k]
                domain_name = info_base['domain']
                domain_type = info_base['natureName']
                domain_licence = info_base['mainLicence']
                website_licence = info_base['serviceLicence']
                domain_status = info_base['limitAccess']
                domain_approve_date = info_base['updateRecordTime']
                domain_owner = info_base['unitName']
                try:
                    domain_content_approved = info_base['contentTypeName']
                    if domain_content_approved == "":
                        domain_content_approved = "无"
                except KeyError:
                    domain_content_approved = "无"
                row_data = domain_owner, domain_name, domain_licence, website_licence, domain_type, domain_content_approved, domain_status, domain_approve_date
                domain_list.append(row_data)
            info_data = {'pageNum': i+2, 'pageSize': '40', 'unitName': info, "serviceType":"1"}
            if beian_info['params']['isLastPage'] is True:
                break
            else:
                beian_info = requests.post(info_url, json=info_data, headers=base_header).json()
                end_row = beian_info['params']['endRow']
                time.sleep(3)
    except:
        return domain_list
    return domain_list


def data_saver(domain_list,file_path):
    """
    打印最终结果，并保存数据至Excel表格，同时调整表格格式。
    """
    # 计算需要写入表格的总行数，如果是空列表，即代表该域名没有备案信息，也有可能是获取信息失败了
    total_row = len(domain_list)
    if total_row == 1:
        total_row = 0
    elif total_row == 0:
        return print("[*]查询结果：所查域名无备案\n----------------------------------------------------")
    print(f"[*]查询结果：{domain_list}")

    # 存在对应文件，则读取表格追加写入，不存在则创建，并设置表格的标题、列宽、冻结窗格、文字布局等格式
    if os.path.exists(file_path):
        wb = xl.load_workbook(file_path)
        ws = wb['备案信息']
        max_row = ws.max_row
        start = max_row+1
        total_row = total_row + start
        after_title = 0
    else:
        wb = xl.Workbook()
        ws = wb.active
        ws.title = "备案信息"
        title_list = ['域名主办方', '域名', '备案许可证号', '网站备案号', '域名类型', '网站前置审批项', '是否限制接入', '审核通过日期']
        for i in range(0, 8):
            ws.cell(1, i+1).value=title_list[i]
        col_width = {'A': 45, 'B': 40, 'C': 22, 'D': 24, 'E': 9, 'F': 15, 'G': 13, 'H': 21}
        for k, v in col_width.items():
            ws.column_dimensions[k].width = v
        ws.freeze_panes = 'A2'
        start = 0
        after_title = 2
    # 写入查询数据
    for j in range(start, total_row+1):
        for k in range(0, 8):
            try:
                ws.cell(j+after_title, k+1).value = domain_list[j-start][k]
            except:
               continue
    # 垂直居中
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            ws.cell(row+1, col+1).alignment = Alignment(horizontal='center', vertical='center')
    try:
        wb.save(file_path)
    except PermissionError:
        print("** 备案信息登记表格已打开，无法写入文件。如需写入，请关闭文件后重新执行！ **\n")
        return -1
    print(f"[*]保存结果：已保存在：{file_path}")
    print('----------------------------------------------------')
    return 'OK'


def main(infos,file_name_path):
    cookie = get_cookies()
    info = query_base(infos)
    file_path = file_name_path

    try:
        global base_header
        base_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32',
            'Origin': 'https://beian.miit.gov.cn',
            'Referer': 'https://beian.miit.gov.cn/',
            'Cookie': f'__jsluid_s={cookie}'
        }
        # -1代表对应步骤失败了，不是-1则正常执行下一步
        if cookie != -1:
            token = get_token()
            if token != -1:
                check_data = get_check_pic(token)
                if check_data != -1:
                    sign = get_sign(check_data, token)
                    p_uuid = check_data['key']
                    if sign != -1:
                        domain_list = get_beian_info(info, p_uuid, token, sign)
                        data_saver(domain_list,file_path)
                    else:
                        raise ValueError("获取Sign错误，请重试！")
                else:
                    raise ValueError("计算图片缺口位置错误，请重试！")
            else:
                raise ValueError("获取Token失败，如频繁失败请关闭程序后等待几分钟再试！")
        else:
            cookie = get_cookies()
            raise ValueError("获取Cookie失败，请重试！")
    except Exception as e:
        print(f'{e}\n')


if __name__ == '__main__':
    # 创建解析器对象
    parser = argparse.ArgumentParser(description='批量查询ICP备案信息')

    # 创建互斥参数组
    group = parser.add_mutually_exclusive_group()

    # 添加命令行参数选项
    group.add_argument('-o', '--file-path', help='批量查询，输入文件路径')
    group.add_argument('-d', '--doma_name', help='单个查询，输入公司名称或主域名')

    # 解析命令行参数
    args = parser.parse_args()

    # 获取文件路径参数的值
    file_path = args.file_path

    # 获取单个数值
    doma_name = args.doma_name

    # 规定唯一的表格文件路径
    current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_name_path = f"./results/备案信息_{current_time}.xlsx"

    if file_path:
        print('----------------------------------------------------')
        print('[*]读取到传入文件：'+file_path)
        print('[*]数据保存地址：'+file_name_path)
        print('----------------------------------------------------')
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            for line in file:
                line = line.strip()
                main(line,file_name_path)
                # 睡眠5秒，防止接口被绊掉
                time.sleep(5)
    elif doma_name:
        print('----------------------------------------------------')
        main(doma_name,file_name_path)
    else:
        print('请提供文件路径参数。使用 -h 或 --help 选项获取帮助信息。')
